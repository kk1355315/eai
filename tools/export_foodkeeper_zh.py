from __future__ import annotations

import argparse
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Callable
from xml.sax.saxutils import escape

import requests

SHEET_NAME_ZH = {
    "Version": "版本记录",
    "Category": "分类",
    "Product": "食品保质期",
    "CookingTips": "烹饪提示",
    "CookingMethods": "烹饪方法",
    "Data Dictionary": "字段字典",
}

BASE_COLUMN_NAME_ZH = {
    "ID": "编号",
    "Category_ID": "分类编号",
    "Product_ID": "食品编号",
    "Name": "名称",
    "Name_subtitle": "名称副标题",
    "Keywords": "关键词",
    "Category_Name": "分类名称",
    "Subcategory_Name": "子分类名称",
    "Sheet": "工作表",
    "Column": "字段",
    "Description": "说明",
    "Tips": "提示",
    "Safe_Minimum_Temperature": "最低安全温度",
    "Rest_Time": "静置时间",
    "Rest_Time_metric": "静置时间单位",
    "Cooking_Method": "烹饪方式",
    "Measure_from": "规格起始值",
    "Measure_to": "规格结束值",
    "Size_metric": "规格单位",
    "Cooking_Temperature": "烹饪温度",
    "Timing_from": "时间起始值",
    "Timing_to": "时间结束值",
    "Timing_metric": "时间单位",
    "Timing_per": "计时基准",
    "Data_Version_Number": "数据版本号",
    "Current_Version": "当前版本",
    "Modified_Date": "修改日期",
    "FSIS_Approved_Flag": "FSIS批准标记",
    "Approved_Date": "批准日期",
    "Notes": "备注",
}

STORAGE_PREFIX_ZH = {
    "Refrigerate_After_Opening": "开封后冷藏",
    "Refrigerate_After_Thawing": "解冻后冷藏",
    "Pantry_After_Opening": "开封后常温",
    "DOP_Refrigerate": "购买后冷藏",
    "DOP_Pantry": "购买后常温",
    "DOP_Freeze": "购买后冷冻",
    "Refrigerate": "冷藏",
    "Pantry": "常温",
    "Freeze": "冷冻",
}

STORAGE_SUFFIX_ZH = {
    "Min": "最短值",
    "Max": "最长值",
    "Metric": "单位",
    "tips": "提示",
    "Tips": "提示",
}

ENUM_VALUE_ZH = {
    "Days": "天",
    "Weeks": "周",
    "Months": "月",
    "Year": "年",
    "Years": "年",
    "Hours": "小时",
    "When Ripe": "成熟时",
    "Indefinitely": "无限期",
    "Package use-by date": "按包装食用期限",
    "Not Recommended": "不建议",
    "Integer": "整数",
    "Text": "文本",
    "Yes": "是",
    "Y": "是",
    "No": "否",
    "N": "否",
    "minutes": "分钟",
    "minute": "分钟",
    "pound": "磅",
    "pounds": "磅",
}

PHRASE_OVERRIDES = {
    "Baby Food": "婴儿食品",
    "Baked Goods": "烘焙食品",
    "Bakery": "面包糕点",
    "Baking and Cooking": "烘焙与烹调用品",
    "Refrigerated Dough": "冷藏面团",
    "Beverages": "饮料",
    "Condiments, Sauces & Canned Goods": "调味品、酱料和罐头食品",
    "Dairy Products & Eggs": "乳制品和蛋类",
    "Food Purchased Frozen": "购买时即为冷冻的食品",
    "Grains, Beans & Pasta": "谷物、豆类和意面",
    "Meat": "肉类",
    "Fresh": "新鲜",
    "Shelf Stable Foods": "常温耐储食品",
    "Smoked or Processed": "烟熏或加工类",
    "Stuffed or Assembled": "填馅或组合类",
    "Poultry": "禽类",
    "Cooked or Processed": "熟制或加工类",
    "Produce": "农产品",
    "Fresh Fruits": "新鲜水果",
    "Fresh Vegetables": "新鲜蔬菜",
    "Seafood": "海鲜",
    "Shellfish": "贝类海鲜",
    "Smoked": "烟熏类",
    "Vegetarian Proteins": "植物蛋白食品",
    "Deli & Prepared Foods": "熟食与即食食品",
    "Version": "版本记录",
    "Category": "分类",
    "Product": "食品保质期",
    "CookingTips": "烹饪提示",
    "CookingMethods": "烹饪方法",
    "Data Dictionary": "字段字典",
}

TEXT_TRANSLATION_FIELDS = {
    "Name",
    "Name_subtitle",
    "Category_Name",
    "Subcategory_Name",
    "Tips",
    "Pantry_tips",
    "DOP_Pantry_tips",
    "Refrigerate_tips",
    "DOP_Refrigerate_tips",
    "Freeze_Tips",
    "DOP_Freeze_Tips",
    "Notes",
    "Cooking_Method",
    "Timing_per",
    "Size_metric",
}

FAST_TEXT_TRANSLATION_FIELDS = {
    "Name",
    "Category_Name",
    "Subcategory_Name",
    "Cooking_Method",
}


def flatten_sheet_rows(sheet: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in sheet.get("data", []):
        flat_row: dict[str, Any] = {}
        if isinstance(row, list):
            for cell in row:
                if isinstance(cell, dict):
                    flat_row.update(cell)
        elif isinstance(row, dict):
            flat_row.update(row)
        rows.append(flat_row)
    return rows


def translate_sheet_name(name: str) -> str:
    return SHEET_NAME_ZH.get(name, name)


def translate_column_name(name: str) -> str:
    if name in BASE_COLUMN_NAME_ZH:
        return BASE_COLUMN_NAME_ZH[name]

    for prefix, prefix_zh in sorted(STORAGE_PREFIX_ZH.items(), key=lambda item: -len(item[0])):
        for suffix, suffix_zh in STORAGE_SUFFIX_ZH.items():
            if name == f"{prefix}_{suffix}":
                return f"{prefix_zh}_{suffix_zh}"

    return name


def combine_translation(original: str, translated: str) -> str:
    normalized_original = original.strip()
    normalized_translated = translated.strip()
    if not normalized_translated or normalized_translated == normalized_original:
        return normalized_original
    return f"{normalized_translated} | {normalized_original}"


def translate_enum(value: str) -> str:
    return ENUM_VALUE_ZH.get(value, value)


def translate_description_text(value: str, translator: Callable[[str], str]) -> str:
    if not value:
        return value

    parts = [part.strip() for part in value.split(",")]
    if len(parts) > 1:
        translated_parts = [
            translate_enum(part) if translate_enum(part) != part else translator(part).strip()
            for part in parts
        ]
        return ", ".join(translated_parts)

    translated = translate_enum(value)
    if translated != value:
        return translated

    translated = translator(value)
    return translated.strip() if translated else value


class CachedTranslator:
    def __init__(
        self,
        backend: Callable[[str], str] | None = None,
        preloaded_cache: dict[str, str] | None = None,
    ) -> None:
        self._backend = backend or (lambda text: text)
        self._cache: dict[str, str] = dict(preloaded_cache or {})

    def __call__(self, text: str) -> str:
        normalized = text.strip()
        if not normalized:
            return text
        if normalized in self._cache:
            return self._cache[normalized]
        try:
            translated = self._backend(normalized)
        except Exception:
            translated = normalized
        result = translated.strip() if translated else normalized
        self._cache[normalized] = result
        return result


def create_default_backend() -> Callable[[str], str]:
    def translate(text: str) -> str:
        response = requests.get(
            "https://translate.googleapis.com/translate_a/single",
            params={
                "client": "gtx",
                "sl": "en",
                "tl": "zh-CN",
                "dt": "t",
                "q": text,
            },
            timeout=8,
        )
        response.raise_for_status()
        payload = response.json()
        return "".join(part[0] for part in payload[0])

    return translate


def translate_phrase(text: str, translator: Callable[[str], str]) -> str:
    if text in PHRASE_OVERRIDES:
        return PHRASE_OVERRIDES[text]
    translated_enum = translate_enum(text)
    if translated_enum != text:
        return translated_enum
    return translator(text)


def collect_unique_texts(workbook: dict[str, Any], translatable_fields: set[str]) -> list[str]:
    texts: list[str] = []
    seen: set[str] = set()
    for sheet in workbook.get("sheets", []):
        for row in flatten_sheet_rows(sheet):
            for column_name, value in row.items():
                if (
                    column_name in translatable_fields
                    and isinstance(value, str)
                    and value.strip()
                    and value.strip() not in seen
                    and value.strip() not in PHRASE_OVERRIDES
                    and translate_enum(value.strip()) == value.strip()
                ):
                    seen.add(value.strip())
                    texts.append(value.strip())
    return texts


def pretranslate_texts(
    texts: list[str],
    backend: Callable[[str], str],
    max_workers: int = 12,
) -> dict[str, str]:
    if not texts:
        return {}

    cache: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(backend, text): text for text in texts}
        for future in as_completed(future_map):
            text = future_map[future]
            try:
                translated = future.result()
            except Exception:
                translated = text
            cache[text] = translated.strip() if translated else text
    return cache


def translate_value(
    column_name: str,
    value: Any,
    translator: Callable[[str], str],
    translatable_fields: set[str],
) -> Any:
    if value is None:
        return None

    if isinstance(value, bool):
        return "是" if value else "否"

    if isinstance(value, (int, float)):
        return value

    if not isinstance(value, str):
        return value

    if column_name == "Sheet":
        return combine_translation(value, translate_sheet_name(value))

    if column_name == "Column":
        return combine_translation(value, translate_column_name(value))

    if column_name == "Description":
        return translate_description_text(value, translator)

    if column_name == "Keywords":
        return value

    if column_name in translatable_fields:
        return combine_translation(value, translate_phrase(value, translator))

    translated_enum = translate_enum(value)
    if translated_enum != value:
        return translated_enum

    return value


def build_translated_workbook(
    workbook: dict[str, Any],
    translator: Callable[[str], str] | None = None,
    translatable_fields: set[str] | None = None,
) -> dict[str, Any]:
    active_fields = translatable_fields or TEXT_TRANSLATION_FIELDS
    backend = translator or create_default_backend()
    preload = {} if translator is not None else pretranslate_texts(
        collect_unique_texts(workbook, active_fields),
        backend,
    )
    cached_translator = CachedTranslator(backend, preloaded_cache=preload)
    translated_sheets: list[dict[str, Any]] = []

    for sheet in workbook.get("sheets", []):
        rows = flatten_sheet_rows(sheet)
        translated_rows: list[dict[str, Any]] = []
        for row in rows:
            translated_row: dict[str, Any] = {}
            for column_name, value in row.items():
                translated_row[translate_column_name(column_name)] = translate_value(
                    column_name,
                    value,
                    cached_translator,
                    active_fields,
                )
            translated_rows.append(translated_row)

        translated_sheets.append(
            {
                "name": translate_sheet_name(sheet["name"]),
                "rows": translated_rows,
            }
        )

    return {"fileName": workbook.get("fileName"), "sheets": translated_sheets}


def xml_cell(value: Any) -> str:
    if value is None:
        return "<Cell/>"

    if isinstance(value, bool):
        data_type = "String"
        serialized = "是" if value else "否"
    elif isinstance(value, (int, float)):
        data_type = "Number"
        serialized = str(int(value) if isinstance(value, float) and value.is_integer() else value)
    else:
        data_type = "String"
        serialized = str(value)

    return f'<Cell><Data ss:Type="{data_type}">{escape(serialized)}</Data></Cell>'


def write_excel_xml(workbook: dict[str, Any], output_path: Path) -> None:
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<?mso-application progid="Excel.Sheet"?>',
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"',
        ' xmlns:o="urn:schemas-microsoft-com:office:office"',
        ' xmlns:x="urn:schemas-microsoft-com:office:excel"',
        ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"',
        ' xmlns:html="http://www.w3.org/TR/REC-html40">',
    ]

    for sheet in workbook.get("sheets", []):
        rows = sheet.get("rows", [])
        headers = list(rows[0].keys()) if rows else []
        lines.append(f'<Worksheet ss:Name="{escape(sheet["name"])}">')
        lines.append("<Table>")
        if headers:
            lines.append("<Row>")
            for header in headers:
                lines.append(xml_cell(header))
            lines.append("</Row>")
        for row in rows:
            lines.append("<Row>")
            for header in headers:
                lines.append(xml_cell(row.get(header)))
            lines.append("</Row>")
        lines.append("</Table>")
        lines.append("</Worksheet>")

    lines.append("</Workbook>")
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(workbook: dict[str, Any], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to write .xlsx output") from exc

    excel = Workbook()
    default_sheet = excel.active
    excel.remove(default_sheet)

    for sheet in workbook.get("sheets", []):
        rows = sheet.get("rows", [])
        headers = list(rows[0].keys()) if rows else []
        worksheet = excel.create_sheet(title=sheet["name"][:31] or "Sheet1")

        if headers:
            worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header) for header in headers])

    excel.save(output_path)


def load_workbook(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Translate foodkeeper.json into a Chinese Excel-compatible workbook."
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=Path("foodkeeper.json"),
        help="Path to foodkeeper.json",
    )
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=Path("foodkeeper_zh.xml"),
        help="Output Excel XML workbook path",
    )
    parser.add_argument(
        "--full-text",
        action="store_true",
        help="Translate longer free-text fields as well. Slower on large files.",
    )
    args = parser.parse_args()

    workbook = load_workbook(args.input)
    translatable_fields = TEXT_TRANSLATION_FIELDS if args.full_text else FAST_TEXT_TRANSLATION_FIELDS
    translated = build_translated_workbook(workbook, translatable_fields=translatable_fields)
    if args.output.suffix.lower() == ".xlsx":
        write_xlsx(translated, args.output)
    else:
        write_excel_xml(translated, args.output)
    print(f"wrote translated workbook to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
